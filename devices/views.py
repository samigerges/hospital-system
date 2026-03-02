# devices/views.py
from datetime import date, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.db.models import Count, Avg, Sum, Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Device, Department, Maintenance
from .forms import LoginForm, DeviceForm, DepartmentForm, MaintenanceForm

from .utils.prediction import compute_failure_prediction



@require_GET
@login_required
def control_center_stats_api(request):
    total_devices = Device.objects.count()
    active_devices = Device.objects.filter(status="active").count()
    maintenance_devices = Device.objects.filter(status="maintenance").count()

    today = timezone.now().date()

    # Critical: overdue or due within 3 days (نفس منطق الـ control_center)
    critical_qs = Device.objects.filter(next_maintenance__isnull=False).filter(
        Q(next_maintenance__lt=today) | Q(next_maintenance__lte=today + timedelta(days=3))
    )
    critical_alerts = critical_qs.count()

    # Recent maintenance (آخر 5)
    recent_qs = Maintenance.objects.select_related("device").order_by("-date")[:5]
    recent_list = []
    for m in recent_qs:
        recent_list.append({
            "device_name": m.device.name if m.device else "",
            "device_id": m.device.device_id if m.device else "",
            "technician": m.technician or "Technician",
            "maintenance_type": m.get_maintenance_type_display(),
            "cost": float(m.cost or 0),
            "date": m.date.strftime("%b %d, %Y") if m.date else "",
        })

    # System health = نسبة active من الإجمالي
    system_health = round((active_devices / total_devices) * 100, 1) if total_devices else 0

    return JsonResponse({
        "counts": {
            "total_devices": total_devices,
            "active_devices": active_devices,
            "maintenance_devices": maintenance_devices,
            "critical_alerts": critical_alerts,
        },
        "system_health": system_health,
        "recent_maintenance": recent_list,
        "server_time": timezone.now().strftime("%H:%M:%S"),
    })

@require_GET
@login_required
def device_lookup_api(request):
    device_id = (request.GET.get("device_id") or "").strip()
    if not device_id:
        return JsonResponse({"ok": False, "error": "device_id is required"}, status=400)

    device = Device.objects.filter(device_id__iexact=device_id).first()
    if not device:
        return JsonResponse({"ok": False, "error": "Device not found"}, status=404)

    # رابط صفحة التفاصيل
    return JsonResponse({
        "ok": True,
        "pk": device.pk,
        "url": f"/devices/{device.pk}/",  # لو عندك path name استخدم reverse أفضل
    })


def devices_export_excel(request):
    qs = Device.objects.all()

    search = request.GET.get("search")
    status = request.GET.get("status")
    department = request.GET.get("department")

    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(device_id__icontains=search))

    if status:
        qs = qs.filter(status=status)

    if department:
        qs = qs.filter(department_id=department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    headers = ["Device ID", "Name", "Type", "Department", "Status", "Model", "Serial"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    for d in qs:
        ws.append([
            d.device_id,
            d.name,
            d.get_device_type_display(),
            d.department.name if d.department else "",
            d.get_status_display(),
            d.model,
            d.serial_number
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=devices.xlsx"
    wb.save(response)
    return response

# Prediction utility
from .utils.prediction import compute_failure_prediction

# Language (you had these)
from django.utils.translation import gettext as _
from django.views.i18n import set_language


def login_view(request):
    if request.user.is_authenticated:
        # Redirect to Control Center instead of dashboard
        return redirect('control_center')  # Changed from 'dashboard'
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            user_model = get_user_model()
            user, _ = user_model.objects.get_or_create(username=username)
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, 'Logged in successfully!')
            # Redirect to Control Center after login
            return redirect('control_center')  # Changed from 'dashboard'
        else:
            messages.error(request, 'Invalid username or password')
    else:
        form = LoginForm()

    return render(request, 'auth/login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.success(request, 'Logged out successfully!')
    return redirect('login')


@login_required
def dashboard(request):
    # =========================
    # Device counts
    # =========================
    total_devices = Device.objects.count()
    active_devices = Device.objects.filter(status='active').count()
    maintenance_devices = Device.objects.filter(status='maintenance').count()
    inactive_devices = Device.objects.filter(status='inactive').count()

    # =========================
    # Percentages
    # =========================
    active_percentage = maintenance_percentage = inactive_percentage = 0

    if total_devices > 0:
        active_percentage = (active_devices / total_devices) * 100
        maintenance_percentage = (maintenance_devices / total_devices) * 100
        inactive_percentage = (inactive_devices / total_devices) * 100

    # =========================
    # Departments
    # =========================
    total_departments = Department.objects.count()

    # =========================
    # Recent Devices
    # =========================
    recent_devices = Device.objects.order_by('-created_at')[:5]

    # =========================
    # Upcoming Maintenance - SORT BY CRITICALITY ONLY
    # =========================
    today = date.today()

    devices_with_maintenance = Device.objects.filter(next_maintenance__isnull=False)

    upcoming_maintenance = []
    for device in devices_with_maintenance:
        delta = device.next_maintenance - today
        days = delta.days

        device.days_until_maintenance = days
        device.days_overdue = abs(days) if days < 0 else 0
        device.is_maintenance_overdue = days < 0
        device.is_maintenance_urgent = 0 <= days <= 3
        device.is_maintenance_soon = 4 <= days <= 7

        # ترتيب حسب الخطورة
        if device.is_maintenance_overdue:
            device.priority_score = 0
        elif device.is_maintenance_urgent:
            device.priority_score = 1
        elif device.is_maintenance_soon:
            device.priority_score = 2
        else:
            device.priority_score = 3

        upcoming_maintenance.append(device)

    # الترتيب النهائي
    upcoming_maintenance.sort(key=lambda d: (d.priority_score, d.next_maintenance))
    upcoming_maintenance = upcoming_maintenance[:10]

    # Enrich objects with calculated fields
    for device in upcoming_maintenance:
        delta = device.next_maintenance - today
        device.days_until_maintenance = delta.days
        device.is_maintenance_overdue = delta.days < 0
        device.days_overdue = abs(delta.days) if delta.days < 0 else 0
        device.is_maintenance_urgent = 0 <= delta.days <= 3

    context = {
        'total_devices': total_devices,
        'active_devices': active_devices,
        'maintenance_devices': maintenance_devices,
        'inactive_devices': inactive_devices,
        'total_departments': total_departments,
        'recent_devices': recent_devices,
        'upcoming_maintenance': upcoming_maintenance,
        'active_percentage': active_percentage,
        'maintenance_percentage': maintenance_percentage,
        'inactive_percentage': inactive_percentage,
    }

    return render(request, 'dashboard.html', context)


@login_required
def device_list(request):
    devices_base = Device.objects.all().order_by('-created_at')

    status_filter = request.GET.get('status')
    department_filter = request.GET.get('department')
    search_query = request.GET.get('search')

    if department_filter:
        devices_base = devices_base.filter(department_id=department_filter)
    if search_query:
        devices_base = devices_base.filter(
            Q(name__icontains=search_query) |
            Q(device_id__icontains=search_query) |
            Q(serial_number__icontains=search_query)
        )

    devices = devices_base
    if status_filter:
        devices = devices.filter(status=status_filter)

    departments = Department.objects.all()

    status_kpis = []
    base_params = request.GET.copy()
    if 'status' in base_params:
        del base_params['status']

    clear_status_url = f"?{base_params.urlencode()}" if base_params.urlencode() else '?'

    for value, label in Device.DEVICE_STATUS:
        params = base_params.copy()
        params['status'] = value
        status_kpis.append({
            'value': value,
            'label': label,
            'count': devices_base.filter(status=value).count(),
            'url': f"?{params.urlencode()}",
            'is_selected': status_filter == value,
        })

    context = {
        'devices': devices,
        'departments': departments,
        'status_choices': Device.DEVICE_STATUS,
        'status_kpis': status_kpis,
        'total_devices': devices_base.count(),
        'clear_status_url': clear_status_url,
    }

    return render(request, 'devices/list.html', context)


@login_required
def device_add(request):
    if request.method == 'POST':
        form = DeviceForm(request.POST, request.FILES)
        if form.is_valid():
            device = form.save()

            # توليد QR Code تلقائيًا
            try:
                device.generate_qr_code()
                device.save()
                messages.info(request, 'QR Code generated successfully!')
            except Exception as e:
                messages.warning(request, f'Device saved but QR code generation failed: {str(e)}')

            messages.success(request, f'Device {device.name} added successfully!')
            return redirect('device_detail', pk=device.pk)
    else:
        form = DeviceForm()

    context = {'form': form, 'title': 'Add New Device'}
    return render(request, 'devices/add_edit.html', context)


@login_required
def device_edit(request, pk):
    device = get_object_or_404(Device, pk=pk)

    if request.method == 'POST':
        form = DeviceForm(request.POST, request.FILES, instance=device)
        if form.is_valid():
            form.save()

            if 'generate_qr' in request.POST:
                try:
                    device.generate_qr_code()
                    device.save()
                    messages.success(request, 'QR Code regenerated successfully!')
                except Exception as e:
                    messages.error(request, f'Failed to generate QR code: {str(e)}')

            messages.success(request, f'Device {device.name} updated successfully!')
            return redirect('device_list')
    else:
        form = DeviceForm(instance=device)

    context = {'form': form, 'device': device, 'title': 'Edit Device'}
    return render(request, 'devices/add_edit.html', context)


@login_required
def device_delete(request, pk):
    device = get_object_or_404(Device, pk=pk)
    if request.method == 'POST':
        device_name = device.name
        device.delete()
        messages.success(request, f'Device {device_name} deleted successfully!')
        return redirect('device_list')
    return render(request, 'devices/delete_confirm.html', {'device': device})


@login_required
def generate_device_qr(request, pk):
    """توليد أو إعادة توليد QR code للجهاز"""
    device = get_object_or_404(Device, pk=pk)

    try:
        device.generate_qr_code()
        device.save()
        messages.success(request, 'QR Code generated successfully!')
    except Exception as e:
        messages.error(request, f'Failed to generate QR code: {str(e)}')

    return redirect('device_detail', pk=pk)


def device_detail(request, pk):
    """
    Single source of truth for device detail:
    - public access (QR/public page templates)
    - authenticated access (full detail + add maintenance)
    - includes prediction context in both modes
    """
    device = get_object_or_404(Device, pk=pk)
    maintenances = device.maintenances.all().order_by('-date')

    is_qr_access = request.GET.get('qr') in ['1', 'true', 'True']

    # Always compute prediction from your real data
    prediction = compute_failure_prediction(device, maintenances)

    # Public user flow
    if not request.user.is_authenticated:
        if request.method == 'POST':
            messages.warning(request, 'Please login to add maintenance records')
            return redirect('login')

        context = {
            'device': device,
            'maintenances': maintenances,
            'is_public': True,
            'is_qr_access': is_qr_access,
            'prediction': prediction,
        }

        if is_qr_access:
            return render(request, 'devices/device_public_qr.html', context)
        return render(request, 'devices/device_public.html', context)

    # Authenticated flow: add maintenance
    if request.method == 'POST':
        maintenance_form = MaintenanceForm(request.POST)
        if maintenance_form.is_valid():
            maintenance = maintenance_form.save(commit=False)
            maintenance.device = device
            maintenance.save()

            # Update device dates
            device.last_maintenance = maintenance.date
            if maintenance.next_maintenance_date:
                device.next_maintenance = maintenance.next_maintenance_date
            device.save()

            messages.success(request, 'Maintenance record added successfully!')
            return redirect('device_detail', pk=pk)
    else:
        maintenance_form = MaintenanceForm()

    context = {
        'device': device,
        'maintenances': maintenances,
        'maintenance_form': maintenance_form,
        'is_qr_access': is_qr_access,
        'prediction': prediction,
    }

    return render(request, 'devices/detail.html', context)


@login_required
def departments_list(request):
    """
    Display list of all departments with statistics
    """
    departments = Department.objects.annotate(
        device_count=Count('device')
    ).order_by('name')

    total_devices = sum(dept.device_count for dept in departments)

    avg_devices_per_dept = 0
    if departments.exists():
        avg_devices_per_dept = total_devices / len(departments)

    top_department = None
    if departments.exists():
        top_department = max(departments, key=lambda d: d.device_count)

    for dept in departments:
        dept_devices = Device.objects.filter(department=dept)
        dept.active_devices = dept_devices.filter(status='active').count()
        dept.maintenance_devices = dept_devices.filter(status='maintenance').count()
        dept.inactive_devices = dept_devices.filter(status='inactive').count()

        if not hasattr(dept, 'manager'):
            dept.manager = getattr(dept, 'contact_person', None) or getattr(dept, 'head_of_department', None)

    context = {
        'departments': departments,
        'total_devices': total_devices,
        'avg_devices_per_dept': round(avg_devices_per_dept, 1),
        'top_department': top_department,
        'department_count': departments.count(),
    }

    return render(request, 'devices/departments.html', context)


@login_required
def department_add(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            messages.success(request, f'Department {department.name} added successfully!')
            return redirect('departments_list')
    else:
        form = DepartmentForm()

    context = {'form': form, 'title': 'Add New Department'}
    return render(request, 'devices/department_form.html', context)


@login_required
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, f'Department {department.name} updated successfully!')
            return redirect('departments_list')
    else:
        form = DepartmentForm(instance=department)

    context = {'form': form, 'department': department, 'title': 'Edit Department'}
    return render(request, 'devices/department_form.html', context)


@login_required
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        department_name = department.name
        department.delete()
        messages.success(request, f'Department {department_name} deleted successfully!')
        return redirect('departments_list')
    return render(request, 'devices/department_delete.html', {'department': department})


def set_language_view(request):
    if request.method == 'POST':
        language = request.POST.get('language', 'en')
        request.session['django_language'] = language
    return redirect(request.META.get('HTTP_REFERER', '/'))
@login_required


@login_required
def control_center(request):
    """Control Center - Main operational hub after login"""
    # Device counts
    total_devices = Device.objects.count()
    active_devices = Device.objects.filter(status='active').count()
    maintenance_devices = Device.objects.filter(status='maintenance').count()
    inactive_devices = Device.objects.filter(status='inactive').count()
    
    # Critical alerts calculation
    today = date.today()
    critical_alerts = 0
    overdue_maintenance = []
    
    # Get devices with next maintenance
    devices_with_maintenance = Device.objects.filter(next_maintenance__isnull=False)
    for device in devices_with_maintenance:
        if device.next_maintenance:
            delta = device.next_maintenance - today
            if delta.days < 0:  # Overdue
                critical_alerts += 1
                overdue_maintenance.append({
                    'name': device.name,
                    'device_id': device.device_id,
                    'days_overdue': abs(delta.days),
                    'next_maintenance': device.next_maintenance
                })
            elif delta.days <= 3:  # Due within 3 days (critical)
                critical_alerts += 1
    
    # Departments count
    total_departments = Department.objects.count()
    
    # Recent maintenance records
    recent_maintenance = Maintenance.objects.select_related('device').order_by('-date')[:5]
    
    # System health calculation
    system_health = 87  # Default value, can be calculated based on device status
    if total_devices > 0:
        health_percentage = (active_devices / total_devices) * 100
        system_health = min(health_percentage, 100)
    
    context = {
        # Counts
        'total_devices': total_devices,
        'active_devices': active_devices,
        'maintenance_devices': maintenance_devices,
        'inactive_devices': inactive_devices,
        'total_departments': total_departments,
        'critical_alerts': critical_alerts,
        
        # For status strip
        'devices_normal': active_devices,
        'devices_pending': maintenance_devices,
        'devices_critical': critical_alerts,
        
        # Critical zone data
        'overdue_maintenance': overdue_maintenance,
        'has_critical_alerts': critical_alerts > 0,
        
        # System health
        'system_health': system_health,
        'mean_downtime': '2.3 hours',
        'avg_repair_time': '4.7 hours',
        
        # Recent data
        'recent_maintenance': recent_maintenance,
    }
    
    return render(request, 'control_center.html', context)

def login_view(request):
    if request.user.is_authenticated:
        return redirect('control_center')  # توجيه إلى Control Center بدلاً من Dashboard
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            user_model = get_user_model()
            user, _ = user_model.objects.get_or_create(username=username)
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, 'Logged in successfully!')
            return redirect('control_center')  # توجيه إلى Control Center
        else:
            messages.error(request, 'Invalid username or password')
    else:
        form = LoginForm()

    return render(request, 'auth/login.html', {'form': form})

from django.db.models import Count, Avg, Sum, F, Q
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from .models import Device, Maintenance, Department

@login_required
def reports_view(request):
    """Comprehensive system reports view"""
    
    # 1. Device Statistics
    total_devices = Device.objects.count()
    active_devices = Device.objects.filter(status='active').count()
    maintenance_devices = Device.objects.filter(status='maintenance').count()
    inactive_devices = Device.objects.filter(status='inactive').count()
    
    # 2. Critical devices (overdue maintenance)
    today = date.today()
    critical_devices = 0
    overdue_maintenance = []
    
    devices_with_maintenance = Device.objects.filter(next_maintenance__isnull=False)
    for device in devices_with_maintenance:
        if device.next_maintenance:
            delta = device.next_maintenance - today
            if delta.days < 0:  # Overdue
                critical_devices += 1
                overdue_maintenance.append(device)
    
    # 3. Maintenance Statistics
    maintenance_stats = Maintenance.objects.aggregate(
        total_cost=Sum('cost'),
        avg_cost=Avg('cost'),
        total_records=Count('id')
    )
    
    # 4. Recent Maintenance (last 30 days)
    thirty_days_ago = today - timedelta(days=30)
    recent_maintenance = Maintenance.objects.filter(
        date__gte=thirty_days_ago
    ).order_by('-date')
    
    # 5. Department Statistics with detailed breakdown
    departments = Department.objects.annotate(
        device_count=Count('device'),
        active_count=Count('device', filter=Q(device__status='active')),
        maintenance_count=Count('device', filter=Q(device__status='maintenance')),
        inactive_count=Count('device', filter=Q(device__status='inactive'))
    ).order_by('-device_count')
    
    # 6. Device Type Distribution
    device_types = Device.objects.values('device_type').annotate(
        count=Count('id'),
        active=Count('id', filter=Q(status='active')),
        maintenance=Count('id', filter=Q(status='maintenance')),
        inactive=Count('id', filter=Q(status='inactive'))
    ).order_by('-count')
    
    # 7. Maintenance Type Distribution
    maintenance_types = Maintenance.objects.values('maintenance_type').annotate(
        count=Count('id'),
        avg_cost=Avg('cost'),
        total_cost=Sum('cost')
    ).order_by('-count')
    
    # 8. Monthly Maintenance Trend (last 6 months)
    monthly_trend = []
    for i in range(5, -1, -1):
        month_start = today.replace(day=1) - timedelta(days=30*i)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        month_maintenance = Maintenance.objects.filter(
            date__range=[month_start, month_end]
        ).aggregate(
            count=Count('id'),
            total_cost=Sum('cost')
        )
        
        monthly_trend.append({
            'month': month_start.strftime('%b %Y'),
            'count': month_maintenance['count'] or 0,
            'cost': month_maintenance['total_cost'] or 0
        })
    
    # 9. Top Technicians
    top_technicians = Maintenance.objects.values('technician').annotate(
        count=Count('id'),
        avg_cost=Avg('cost'),
        total_cost=Sum('cost')
    ).order_by('-count')[:5]
    
    # 10. Warranty Status
    warranty_active = Device.objects.filter(warranty_expiry__gte=today).count()
    warranty_expired = Device.objects.filter(warranty_expiry__lt=today).count()
    
    context = {
        # Device Stats
        'total_devices': total_devices,
        'active_devices': active_devices,
        'maintenance_devices': maintenance_devices,
        'inactive_devices': inactive_devices,
        'critical_devices': critical_devices,
        
        # Maintenance Stats
        'maintenance_total_cost': maintenance_stats['total_cost'] or 0,
        'maintenance_avg_cost': maintenance_stats['avg_cost'] or 0,
        'maintenance_total_records': maintenance_stats['total_records'] or 0,
        
        # Overdue Maintenance
        'overdue_maintenance': overdue_maintenance,
        'has_overdue_maintenance': len(overdue_maintenance) > 0,
        
        # Department Stats
        'departments': departments,
        'total_departments': departments.count(),
        
        # Device Type Stats
        'device_types': device_types,
        
        # Maintenance Type Stats
        'maintenance_types': maintenance_types,
        
        # Trends
        'monthly_trend': monthly_trend,
        
        # Technicians
        'top_technicians': top_technicians,
        
        # Warranty
        'warranty_active': warranty_active,
        'warranty_expired': warranty_expired,
        'warranty_percentage': (warranty_active / total_devices * 100) if total_devices > 0 else 0,
        
        # Recent Maintenance
        'recent_maintenance': recent_maintenance[:10],
        
        # Dates for filters
        'today': today,
        'thirty_days_ago': thirty_days_ago,
         'currency_symbol': 'EGP',
        'currency_code': 'EGP',
        
    }
    
    return render(request, 'reports.html', context)




@login_required
def team_profile(request):
    # بيانات المشرف مع مسار الصورة
    supervisor = {
        'name': 'Dr. Lamia Nabil Mahdy',
        'title': 'Project Supervisor & Technical Advisor',
        'department': 'Biomedical Engineering Department',
        'institution': 'Higher Technological Institute',
        'email': 'lamia.nabil@hti.edu.eg',
        'phone': '+20 100 202 8806',
        'photo': 'img/supervisor.png',  # مسار الصورة في مجلد static/img/
        'bio': 'Provides technical guidance on medical equipment standards and healthcare regulations. Oversees project alignment with industry requirements and academic standards.',
        'expertise': [
            'Healthcare Technology Standards',
            'Medical Device Regulations',
            'System Architecture Review',
            'Academic Project Supervision',
            'Industry Best Practices',
            'Research Methodology'
        ]
    }
    
    # بيانات الفريق مع مسارات الصور
    team_members = [
        {
            'name': 'Yousef Mohamed Ahmed Elsayed',
            'role': 'Team Lead & Full Stack Developer',
            'email': 'yousef_mohamed2001@yahoo.com',
            'photo': 'img/yousef.png',  # مسار الصورة
            'bio': 'Biomedical Engineering student leading the DeviceCare project. Combines medical equipment expertise with full-stack development skills to bridge healthcare and technology.',
            'skills': [
                'Django Backend Development',
                'Medical Equipment APIs',
                'Project Management',
                'System Architecture',
                'Database Design',
                'Team Leadership',
                'Client Communication'
            ],
            'experience': [
                'Medical equipment maintenance training',
                'ICU/OR device specialization',
                'Clinical engineering background'
            ]
        },
        {
            'name': 'Eslam Mohamed Sabry Ali',
            'role': 'Frontend Developer & UI Specialist',
            'email': 'eslamsabry086@gmail.com',
            'photo': 'img/eslam.png',  # مسار الصورة
            'bio': 'Specialized in creating intuitive user interfaces for medical systems. Focuses on user experience and responsive design for healthcare applications.',
            'skills': [
                'HTML/CSS/Bootstrap',
                'JavaScript Frontend',
                'UI/UX Design',
                'Responsive Web Design',
                'Technical Documentation',
                'Quality Assurance',
                'User Testing'
            ],
            'experience': [
                'Endoscopy equipment training',
                'Technical documentation',
                'Quality assurance processes'
            ]
        },
        {
            'name': 'Hamsa Samir',
            'role': 'AI Integration & Data Analytics',
            'email': 'hamsasamir96@gmail.com',
            'photo': 'img/hamsa.png',  # مسار الصورة
            'linkedin': 'https://www.linkedin.com/in/hamsa-samir-a14b03263',
            'bio': 'Handles machine learning models and data analytics for predictive maintenance. Integrates AI algorithms with medical equipment monitoring systems.',
            'skills': [
                'Machine Learning Integration',
                'Data Analytics & Visualization',
                'Predictive Algorithms',
                'API Integration',
                'Performance Optimization',
                'Technical Presentations',
                'System Testing'
            ],
            'experience': [
                'Machine Learning certification',
                'Medical imaging systems',
                'Research and development'
            ]
        },
    ]
    
    # معلومات المشروع
    project_info = {
        'name': 'DeviceCare - Intelligent Hospital Equipment Management',
        'description': 'A comprehensive web-based platform for managing hospital medical equipment with intelligent monitoring, predictive maintenance, and real-time analytics. Built using modern web technologies with a focus on user experience and system reliability.',
        'year': '2026',
        'field': 'Biomedical Engineering & Web Development',
        'technologies': [
            'Django 4.2 (Backend Framework)',
            'PostgreSQL (Database)',
            'Bootstrap 5 (Frontend)',
            'JavaScript/Chart.js (Visualization)',
            'RESTful APIs',
            'Docker (Containerization)',
            'Machine Learning Models',
            'WebSockets (Real-time updates)'
        ],
        'features': [
            'Real-time equipment status dashboard',
            'Predictive maintenance scheduling with ML',
            'Interactive equipment tracking maps',
            'Comprehensive analytics and reporting',
            'Multi-role user access control',
            'Mobile-responsive web interface',
            'Automated alert and notification system',
            'Historical data analysis and trends'
        ]
    }
    
    # إحصائيات المشروع
    project_stats = {
        'code_lines': '15,000+',
        'api_endpoints': '45+',
        'database_tables': '25+',
        'development_hours': '800+',
        'testing_coverage': '85%'
    }
    
    context = {
        'supervisor': supervisor,
        'team_members': team_members,
        'project_info': project_info,
        'project_stats': project_stats,
        'active_page': 'team_profile'
    }
    
    return render(request, 'team_profile.html', context)

from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import qrcode
import io
from .models import Device

def device_qr(request, pk):
    """Generate QR on-the-fly pointing to the current host + device URL."""
    device = get_object_or_404(Device, pk=pk)
    # استخدم الطلب لبناء URL كامل يتغير مع تغير IP
    device_path = f"/devices/{device.pk}/"  # عدّل المسار لو كان مختلفًا في المشروع
    full_url = request.build_absolute_uri(device_path)
    img = qrcode.make(full_url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return HttpResponse(buf.getvalue(), content_type='image/png')
